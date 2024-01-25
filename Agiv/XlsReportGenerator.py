# This Python file uses the following encoding: utf-8
import os, pathlib
import openpyxl as opx
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from datetime import datetime

try:
    from .CDBManager import CDBManager
except Exception as ex:
    from CDBManager import CDBManager





#Our working workbook
wb = None

#Flag pour modification format du rapport
flg_last_day_only = False  # Report only the last day measures
flg_last_meas_only = False # Report only the last measures sequence

def set_xls_flg_last_day_only(val):
    global flg_last_day_only
    flg_last_day_only = val

def set_xls_flg_last_meas_only(val):
    global flg_last_meas_only
    flg_last_meas_only = val

def save_XLSreport(fname):
    if wb is not None:
        wb.save(fname)

def _gen_worksheet_header(my_ws, date, giv_id, giv_date, aoip, aoipsn, aoipcal):
    # Add title to the Sheet
    my_ws['B3'] = "Measurements records"
    my_ws.merge_cells('B3:E3')
    cl = my_ws['B3']
    cl.font = Font(size=24)
    cl.alignment = Alignment(horizontal='center')
    my_ws.column_dimensions['B'].width = 20
    my_ws.column_dimensions['C'].width = 20
    my_ws.column_dimensions['D'].width = 20
    my_ws['A5'] = "Records date:"; my_ws['C5'] = date
    my_ws['A6'] = "GIV Id:"; my_ws['C6'] = giv_id
    my_ws['A6'] = "GIV calibration date:"; my_ws['C6'] = giv_date
    my_ws['A7'] = "Calibrator"; my_ws['C7'] = aoip
    my_ws['A8'] = "Serial no:"; my_ws['C8'] = aoipsn
    my_ws['A9'] = "Next calibration:"; my_ws['C9'] = ""


def set_h_merged_cells(ws, start, end, row, val, fontsz=14, color='0000F0FF'):
    """ Horizontal mergin with set of the value, font size, and color """
    ws.merge_cells(start_column=start, end_column=end, start_row=row, end_row=row)
    cl = ws.cell(column=start, row= row, value = val)
    cl.alignment = Alignment(horizontal='center')
    cl.font = Font(size=fontsz)
    cl.fill = PatternFill(fill_type="solid",start_color=color,end_color=color)
    return cl
            



def gen_measures_XLSreport(giv_id, db_object):
    global wb
    first_ws = None
    wb = Workbook()
    db_object.connect()
    db_object.register_giv(giv_id)

    dates = db_object.get_dates_of_measures_for_registrered_Giv( flg_last_day_only )
    if isinstance(dates, str):
        lstr = [dates] # Transform str to list of one str
        dates = lstr
    # Loop for each awailable day: make one sheet by day
    for date in dates:  # Loop on the returned string ( modif pour psycopg2 ) 
        sh_name = date
        my_ws = wb.create_sheet(sh_name)
        if first_ws == None:
            first_ws = wb[sh_name]  # Note the first sheet to show it first
        giv_date = db_object.retrive_giv_last_cal(giv_id)
        print (date)
        #Get registered AOIP at this date  
        (aoip,aoipsn,aoipcal) = db_object.get_aoip_info(date)        
        _gen_worksheet_header(my_ws, date, giv_id, giv_date, aoip, aoipsn, aoipcal)     # Generate header 

        col_ctn = 0     # used for differents sequences

        # Check for all "start measures" for this day and loop on each
        starts_measures = db_object.get_measure_sequences_by_date_for_registered_Giv(date, flg_last_meas_only)
        for (start_key, start_date) in starts_measures:
            row_ctn = 12
            print(f"Sequence {start_key} on {start_date}:")
            # Set colums numbers and letters
            col_ref = 2 + col_ctn
            col_meas =  col_ref + 1
            col_err = col_meas + 1
            lt_ref = get_column_letter(col_ref)
            lt_meas = get_column_letter(col_meas)
            lt_err = get_column_letter(col_err)
            # Display the sequence date
            set_h_merged_cells(my_ws, col_ref, col_err, row_ctn, start_date, 14, '00FFF0FF' )
            row_ctn +=1

            # Set columns size
            for c in range(col_ref,col_meas+1):
                my_ws.column_dimensions[get_column_letter(c)].width = 20
            my_ws.column_dimensions[get_column_letter(c+1)].width = 10 # The column to separate results
            # Loop on all awailable ranges
            ranges_rec = db_object.get_ranges_of_measures_by_date_and_start_for_registrered_Giv(date, start_key)
            for (r_indx, r_name) in ranges_rec:
                print(f"range: {r_name}")
                full_scale = db_object.get_range_fullscale(r_indx)  # Get fullscale value if not null

                row_ctn +=1
                set_h_merged_cells(my_ws, col_ref, col_err, row_ctn, r_name, 14, '0000F0FF' )
                row_ctn += 1

                cl = my_ws.cell(row=row_ctn, column=col_ref, value="Reference").alignment = Alignment(horizontal='center')
                my_ws.cell(row=row_ctn, column=col_meas, value="Measure").alignment = Alignment(horizontal='center')
                my_ws.cell(row=row_ctn, column=col_err, value=f"Error % FS").alignment = Alignment(horizontal='center')
                row_ctn += 1
                meas_recs = db_object.get_measures_by_range_date_giv_id(date, r_indx, start_key)

                # A point record contains: [0]:Date, [1]:Range index, [2]:Ref,  [3]:Measure
                # search the full-scale value to display the % error fct(full_scale)
                if full_scale == 0:
                    for (date_m, range_key, ref, measure, flgko) in meas_recs: 
                        if ref > full_scale:
                            full_scale = ref

                # Display the results for this range 
                for (date_m, range_key, ref, measure, flgko) in meas_recs:
                    cr = my_ws.cell(column=col_ref, row=row_ctn, value = ref)
                    cr.alignment = Alignment(horizontal='center')
                    cm = my_ws.cell(column=col_meas, row=row_ctn, value= measure)
                    cm.alignment = Alignment(horizontal='center')
                    cm.number_format = u'#0.0000'
                    formula = f"=({lt_meas}{row_ctn}-{lt_ref}{row_ctn})/{full_scale}"
                    cerr = my_ws.cell(column=col_err, row=row_ctn, value= formula)
                    cerr.number_format = f"0.000%"
                    cerr.alignment = Alignment(horizontal='center')
                    color = '00FFCC99'
                    if flgko:
                        cerr.fill = PatternFill(fill_type="solid",start_color=color,end_color=color)
                    row_ctn += 1
                row_ctn += 1
            col_ctn += 4
    if first_ws:
        wb.active = first_ws     # Activate the last worsheet
    db_object.close()


def gen_giv_comparaison_report(db_object):
    global wb
    first_ws = None
    wb = Workbook()
    db_object.connect()
    sh_name = "All GIV"
    my_ws = wb.create_sheet(sh_name)

    list_giv = db_object.get_all_givs_and_last_measures_dates()
    col_ctn = 1     # used for differents GIV
    for (key_id, giv_id, date_m) in list_giv:
        print(f"******* GIV {giv_id}  ********")
        db_object.set_default_giv_key(key_id) # Used to retrieve the measures
        row_ctn = 3
        # Set colums numbers and letters
        col_ref = 2 + col_ctn
        col_meas =  col_ref + 1
        col_err = col_meas + 1
        lt_ref = get_column_letter(col_ref)
        lt_meas = get_column_letter(col_meas)
        lt_err = get_column_letter(col_err)
        # Display the GIV ID
        set_h_merged_cells(my_ws, col_ref, col_err, row_ctn, f"GIV ID: {giv_id}", 14, '00FFF0FF' )
        row_ctn +=1
        set_h_merged_cells(my_ws, col_ref, col_err, row_ctn, f"Last measures:{date_m}", 14, '00FFF0FF' )
        row_ctn +=1

        # Set columns size
        for c in range(col_ref,col_meas+1):
            my_ws.column_dimensions[get_column_letter(c)].width = 20
        my_ws.column_dimensions[get_column_letter(c+1)].width = 10 # The column to separate results
        # Loop on all awailable ranges
        #ranges_rec = db_object.get_ranges_of_measure_by_giv_key_for_a_date(key_id, date_m)
        ranges_rec = db_object.get_ranges_of_measures_by_date_and_start_for_registrered_Giv(date_m)
        for (r_indx, r_name) in ranges_rec:
            print(f"range: {r_name}")
            full_scale = db_object.get_range_fullscale(r_indx)  # Get fullscale value if not null
            row_ctn +=1
            set_h_merged_cells(my_ws, col_ref, col_err, row_ctn, r_name, 14, '0000F0FF' )
            row_ctn += 1
            cl = my_ws.cell(row=row_ctn, column=col_ref, value="Reference").alignment = Alignment(horizontal='center')
            my_ws.cell(row=row_ctn, column=col_meas, value="Measure").alignment = Alignment(horizontal='center')
            my_ws.cell(row=row_ctn, column=col_err, value=f"Error % FS").alignment = Alignment(horizontal='center')
            row_ctn += 1
            meas_recs = db_object.get_measures_by_range_date_giv_id(date_m, r_indx)
            # A point record contains: [0]:Date, [1]:Range index, [2]:Ref,  [3]:Measure
            # search the full-scale value to display the % error fct(full_scale)
            if full_scale == 0:
                for (date_t_m, r_indx, ref, measure, flgko) in meas_recs: 
                    if ref > full_scale:
                        full_scale = ref

            # Display the results for this range 
            for (date_t_m, r_indx, ref, measure, flgko) in meas_recs:
                cr = my_ws.cell(column=col_ref, row=row_ctn, value = ref)
                cr.alignment = Alignment(horizontal='center')
                cm = my_ws.cell(column=col_meas, row=row_ctn, value= measure)
                cm.alignment = Alignment(horizontal='center')
                cm.number_format = u'#0.0000'
                formula = f"=({lt_meas}{row_ctn}-{lt_ref}{row_ctn})/{full_scale}"
                cerr = my_ws.cell(column=col_err, row=row_ctn, value= formula)
                cerr.number_format = f"0.000%"
                cerr.alignment = Alignment(horizontal='center')
                color = '00FFCC99'
                if flgko:
                    cerr.fill = PatternFill(fill_type="solid",start_color=color,end_color=color)
                row_ctn += 1
            row_ctn += 1
        col_ctn += 4
    wb.active = my_ws     # Activate the last worsheet
    db_object.close()
    
    pass


# Check excel file generation 
""" Retourne le chemin Agiv dans le home (soit %userprofile%/Agiv sous Windows) """
def get_Agiv_dir():
    home = str(pathlib.Path.home())
    agiv_dir = os.path.join(home,'Agiv')
    return agiv_dir


if __name__ == "__main__":
    giv_id = '850.117'
    my_xls_name = 'report_'+giv_id+".xlsx"
    #my_xls_name = 'report_all_givs.xlsx'
    #db = CDBManager('AP3reports_rec')
    db = CDBManager('etalonnage','utllafond','ELA.AP3saucats', host='10.41.33.97', port='5432')
    #gen_giv_comparaison_report(db)
    gen_measures_XLSreport(giv_id, db)
    curdir = get_Agiv_dir() #os.path.abspath(os.getcwd())
    filename = curdir + "\\" + my_xls_name
    save_XLSreport(filename)
    pass
