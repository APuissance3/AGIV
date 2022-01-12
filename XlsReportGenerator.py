# This Python file uses the following encoding: utf-8
import os
import openpyxl as opx
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from datetime import datetime
from CDBManager import CDBManager



#Our working workbook
wb = None


def save_XLSreport(fname):
    if wb is not None:
        wb.save(fname)

def _gen_worksheet_header(my_ws, date, giv_id, aoip, aoipsn, aoipcal):
    # Add title to the Sheet
    my_ws['B3'] = "Measurements records"
    my_ws.merge_cells('B3:E3')
    cl = my_ws['B3']
    cl.font = Font(size=24)
    cl.alignment = Alignment(horizontal='center')
    my_ws.column_dimensions['B'].width = 20
    my_ws.column_dimensions['C'].width = 20
    my_ws.column_dimensions['D'].width = 20
    my_ws['A5'] = "Records date:"; my_ws['C5'] = date
    my_ws['A6'] = "GIV Id:"; my_ws['C6'] = giv_id
    my_ws['A6'] = "GIV calibration date:"; my_ws['C6'] = "Unknown"
    my_ws['A7'] = "Calibrator"; my_ws['C7'] = aoip
    my_ws['A8'] = "Serial no:"; my_ws['C8'] = aoipsn
    my_ws['A9'] = "Next calibration:"; my_ws['C9'] = ""




def gen_measures_XLSreport(giv_id, db_object):
    global wb
    last_ws = None
    wb = Workbook()
    db_object.connect()
    db_object.register_giv(giv_id)

    dates = db_object.get_dates_of_measures_for_registrered_Giv()

    # Loop for each awailable day: make one sheet by day
    for date in dates:  # Loop on the returned records
        sh_name = "Meas_{}".format(date[0])
        my_ws = wb.create_sheet(sh_name)
        last_ws = wb[sh_name]
        #Get registered AOIP at this date  
        (aoip,aoipsn,aoipcal) = db_object.get_aoip_info(date[0])        
        _gen_worksheet_header(my_ws, date[0], giv_id, aoip, aoipsn, aoipcal)     # Generate header 

        row_ctn = 12
        # Loop on all awailable ranges
        ranges_rec = db_object.get_ranges_of_measures_by_date_for_registrered_Giv(date[0])
        for a_range in ranges_rec:
            row_ctn +=1
            r_name = a_range[1]
            r_indx = a_range[0]
            my_ws.merge_cells(start_column=2, end_column=4, start_row=row_ctn, end_row=row_ctn)
            cl = my_ws.cell(column=2, row= row_ctn)
            cl.value = r_name
            cl.font = Font(size=14)
            cl.alignment = Alignment(horizontal='center')
            cl.fill = PatternFill(fill_type="solid",start_color='1000F0FF',end_color='100000FF')
            row_ctn += 1
            cl = my_ws.cell(row=row_ctn, column=2, value="Reference").alignment = Alignment(horizontal='center')
            my_ws.cell(row=row_ctn, column=3, value="Measure").alignment = Alignment(horizontal='center')
            my_ws.cell(row=row_ctn, column=4, value="Comment").alignment = Alignment(horizontal='center')
            row_ctn += 1
            meas_recs = db_object.get_measures_by_range_date_giv_id(date[0], r_indx)
            # A point record contains: [0]:Date, [1]:Range index, [2]:Ref,  [3]:Measure
            for point_rec in meas_recs:
                cr = my_ws.cell(column=2, row=row_ctn, value=point_rec[2])
                cr.alignment = Alignment(horizontal='center')
                cm = my_ws.cell(column=3, row=row_ctn, value=point_rec[3])
                cm.alignment = Alignment(horizontal='center')
                cm.number_format = u'#0.0000'
                row_ctn += 1
            row_ctn += 1
    wb.active = last_ws     # Activate the last worsheet
    db_object.close()




# Check excel file generation 

if __name__ == "__main__":
    giv_id = '494.647'
    my_xls_name = 'report_'+giv_id+".xlsx"
    db = CDBManager('AP3reports_rec')

    gen_measures_XLSreport(giv_id, db)
    curdir = os.path.abspath(os.getcwd())
    filename = curdir + "\\" + my_xls_name
    save_XLSreport(filename)
